import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# === Load input data ===
base_path = "/home/duch/shipping/speciate/excel/"
speciated_df = pd.read_excel(base_path + "SPECIES.xlsx")
profiles = pd.read_excel(base_path + "PROFILES.xlsx")
species_props = pd.read_excel(base_path + "SPECIES_PROPERTIES.xlsx")

# === Clean types ===
speciated_df['SPECIES_ID'] = speciated_df['SPECIES_ID'].astype(int)
species_props['SPECIES_ID'] = species_props['SPECIES_ID'].astype(int)

# === Filter for shipping-related GAS profiles ===
keywords = ["ship", "marine", "Marine", "vessel", "Vessel", "boat", "barge", "port", "ferry", "Ferry"]
pattern = '|'.join(keywords)
shipping_profiles = profiles[
    profiles['PROFILE_NAME'].str.contains(pattern, case=False, na=False) &
    (profiles['PROFILE_TYPE'].str.upper() == 'GAS')
].sort_values(by='PROFILE_NAME').reset_index(drop=True)

print(f"\nFound {len(shipping_profiles)} candidate shipping-related profiles:\n")
for i, row in shipping_profiles.iterrows():
    print(f"{i}: {row['PROFILE_CODE']} - {row['PROFILE_NAME']}")

# === User selects a profile ===
choice = input("\nEnter the number of the PROFILE_CODE to select: ").strip()
try:
    idx = int(choice)
    profile_code = shipping_profiles.iloc[idx]['PROFILE_CODE']
except (ValueError, IndexError):
    print("❌ Invalid selection. Exiting.")
    exit()

print(f"\nUsing PROFILE_CODE {profile_code} for detailed species fractions...")

# === Extract species fractions ===
fractions = speciated_df[speciated_df['PROFILE_CODE'] == profile_code].copy()
if fractions.empty:
    print(f"No species found for PROFILE_CODE {profile_code}")
    exit()

# === Merge with species names ===
merged = fractions.merge(
    species_props[['SPECIES_ID', 'SPECIES_NAME', 'CAS', 'Molecular Formula', 'SPEC_MW']],
    on='SPECIES_ID',
    how='left'
)

# === Prompt user for total NMVOC emissions ===
nmvoc_total = float(input("\nEnter total NMVOC emission (e.g., in grams): ").strip())
merged['NMVOC_fraction'] = merged['WEIGHT_PERCENT'] / 100
merged['NMVOC_apportioned'] = merged['NMVOC_fraction'] * nmvoc_total

# === Warn if any species have missing names ===
unnamed = merged[merged['SPECIES_NAME'].isna()]
if not unnamed.empty:
    print(f"\n⚠️ {len(unnamed)} species have no SPECIES_NAME in SPECIES_PROPERTIES.xlsx")

# === Sort by emission contribution ===
merged = merged.sort_values(by='NMVOC_apportioned', ascending=False)

# === Show top 15 VOCs ===
print(f"\nTop VOC species for PROFILE_CODE {profile_code} by apportioned NMVOC (g):")
print(merged[['SPECIES_NAME', 'WEIGHT_PERCENT', 'NMVOC_apportioned']].head(15))

# === Plot top 15 as horizontal bar chart ===
top = merged.head(15)
plt.figure(figsize=(10, 6))
plt.barh(top['SPECIES_NAME'], top['NMVOC_apportioned'], color='teal')
plt.xlabel("NMVOC apportioned (g)")
plt.title(f"Top 15 VOC Species for PROFILE_CODE {profile_code}")
plt.gca().invert_yaxis()
plt.tight_layout()
plt.show()

# === Export to Excel with formatting ===
output_excel = f"shipping_profile_{profile_code}_speciated.xlsx"
merged.to_excel(output_excel, index=False)

# Format columns (openpyxl)
wb = load_workbook(output_excel)
ws = wb.active

# Auto-fit column widths
for col in ws.columns:
    max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(10, min(max_length + 2, 40))

# Format NMVOC columns as numbers with 2 decimals
for row in ws.iter_rows(min_row=2, min_col=ws.max_column - 1, max_col=ws.max_column):
    for cell in row:
        cell.number_format = '#,##0.00'

wb.save(output_excel)

print(f"\n✅ Excel file saved: {output_excel}")

